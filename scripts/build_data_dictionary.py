#!/usr/bin/env python3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_fill = PatternFill("solid", fgColor="1B365D")
header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
cell_font = Font(name="Arial", size=10)
thin = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt = PatternFill("solid", fgColor="F4F7FB")


def style_header(ws, cols):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(64+cols)}1"
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_rows(ws, rows):
    for i, row in enumerate(rows, start=2):
        for j, val in enumerate(row, start=1):
            c = ws.cell(i, j, val)
            c.font = cell_font
            c.border = thin
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if i % 2 == 0:
                c.fill = alt


# Inventory
ws = wb.active
ws.title = "Table Inventory"
ws.append(["Layer", "Schema", "Table", "Grain", "Approx rows in sample", "Purpose"])
inv = [
    ["OLTP", "oms", "branch", "1 row per branch", 5, "Advisor office / region"],
    ["OLTP", "oms", "advisor", "1 row per FA", 25, "Financial advisor master"],
    ["OLTP", "oms", "customer", "1 row per client", 120, "Client KYC / risk profile"],
    ["OLTP", "oms", "account", "1 row per brokerage account", 160, "CASH/MARGIN/IRA/401K"],
    ["OLTP", "oms", "exchange", "1 row per venue group", 3, "Listing exchange"],
    ["OLTP", "oms", "security", "1 row per instrument", 20, "Tradable instrument master"],
    ["OLTP", "oms", "app_user", "1 row per login", 25, "Who entered the order"],
    ["OLTP", "oms", "orders", "1 row per order ticket", 15500, "Order lifecycle system of record"],
    ["OLTP", "oms", "execution", "1 row per fill", "~25000", "Partial / full fills"],
    ["OLTP", "oms", "position", "1 row per account+security", "~3000", "Intraday net position"],
    ["OLTP", "oms", "cash_ledger", "1 row per cash posting", "~25000", "Trade cash + fees"],
    ["OLTP", "oms", "order_audit", "1 row per status change", "varies", "Order state machine log"],
    ["OLTP", "oms", "cdc_outbox", "1 row per change event", "varies", "Publish queue to Snowflake"],
    ["Snowflake RAW", "RAW", "ORDERS_CDC", "1 row per CDC event", "varies", "Land JSON payload"],
    ["Snowflake RAW", "RAW", "EXECUTION_CDC", "1 row per CDC event", "varies", "Land JSON payload"],
    ["Snowflake STG", "STG", "ORDERS", "1 current row per order_id", "~15500", "Typed, deduped"],
    ["Snowflake STG", "STG", "EXECUTION", "1 current row per execution_id", "~25000", "Typed, deduped"],
    ["Snowflake MART", "MART", "DIM_DATE", "1 row per calendar day", 3650, "Date role-playing dim"],
    ["Snowflake MART", "MART", "DIM_ADVISOR", "SCD2 advisor", "25+", "Advisor type-2"],
    ["Snowflake MART", "MART", "DIM_CUSTOMER", "SCD2 customer", "120+", "Customer type-2"],
    ["Snowflake MART", "MART", "DIM_ACCOUNT", "SCD2 account", "160+", "Account type-2"],
    ["Snowflake MART", "MART", "DIM_SECURITY", "SCD2 security", "20+", "Instrument type-2"],
    ["Snowflake MART", "MART", "DIM_ORDER_STATUS", "1 row per status", 5, "NEW/PARTIAL/FILLED/CXL/REJ"],
    ["Snowflake MART", "MART", "FACT_ORDER", "1 row per order", "~15500", "Order fact"],
    ["Snowflake MART", "MART", "FACT_EXECUTION", "1 row per fill", "~25000", "Trade/fill fact"],
    ["Snowflake MART", "MART", "FACT_POSITION_SNAPSHOT", "1 row per account+sec+day", "~3000/day", "EOD holdings"],
    ["Snowflake MART", "MART", "AGG_DAILY_VOLUME", "advisor+security+day", "varies", "Dashboard aggregate"],
]
add_rows(ws, inv)
style_header(ws, 6)
ws.column_dimensions["A"].width = 18
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 26
ws.column_dimensions["D"].width = 32
ws.column_dimensions["E"].width = 22
ws.column_dimensions["F"].width = 36

ws2 = wb.create_sheet("Relationships")
ws2.append(["From table", "From col", "To table", "To col", "Cardinality", "Layer"])
rels = [
    ["advisor", "branch_id", "branch", "branch_id", "M:1", "OLTP"],
    ["customer", "advisor_id", "advisor", "advisor_id", "M:1", "OLTP"],
    ["account", "customer_id", "customer", "customer_id", "M:1", "OLTP"],
    ["security", "exchange_id", "exchange", "exchange_id", "M:1", "OLTP"],
    ["orders", "account_id", "account", "account_id", "M:1", "OLTP"],
    ["orders", "security_id", "security", "security_id", "M:1", "OLTP"],
    ["orders", "entered_by", "app_user", "user_id", "M:1", "OLTP"],
    ["execution", "order_id", "orders", "order_id", "M:1", "OLTP"],
    ["execution", "account_id", "account", "account_id", "M:1", "OLTP"],
    ["position", "account_id + security_id", "account / security", "ids", "1:1 current", "OLTP"],
    ["cash_ledger", "execution_id", "execution", "execution_id", "M:1", "OLTP"],
    ["FACT_EXECUTION", "account_sk", "DIM_ACCOUNT", "account_sk", "M:1", "MART"],
    ["FACT_EXECUTION", "security_sk", "DIM_SECURITY", "security_sk", "M:1", "MART"],
    ["FACT_EXECUTION", "trade_date_sk", "DIM_DATE", "date_sk", "M:1", "MART"],
    ["FACT_EXECUTION", "advisor_sk", "DIM_ADVISOR", "advisor_sk", "M:1", "MART"],
    ["FACT_ORDER", "status_sk", "DIM_ORDER_STATUS", "status_sk", "M:1", "MART"],
    ["FACT_POSITION_SNAPSHOT", "as_of_date_sk", "DIM_DATE", "date_sk", "M:1", "MART"],
]
add_rows(ws2, rels)
style_header(ws2, 6)
for col, w in zip("ABCDEF", [22, 26, 22, 16, 16, 10]):
    ws2.column_dimensions[col].width = w

ws3 = wb.create_sheet("FACT_EXECUTION columns")
ws3.append(["Column", "Type", "Source", "Notes"])
cols = [
    ["execution_sk", "NUMBER", "Surrogate", "Warehouse PK"],
    ["execution_id", "NUMBER", "oms.execution", "Natural key from OMS"],
    ["order_id", "NUMBER", "oms.execution", "Parent order"],
    ["account_sk", "NUMBER", "DIM_ACCOUNT", "Current SCD2 key"],
    ["customer_sk", "NUMBER", "DIM_CUSTOMER", "Current SCD2 key"],
    ["advisor_sk", "NUMBER", "DIM_ADVISOR", "Current SCD2 key"],
    ["security_sk", "NUMBER", "DIM_SECURITY", "Current SCD2 key"],
    ["trade_date_sk", "NUMBER", "DIM_DATE", "YYYYMMDD from fill_ts"],
    ["side", "STRING", "oms.orders.side", "B or S"],
    ["fill_qty", "NUMBER(18,4)", "oms.execution", "Filled quantity"],
    ["fill_px", "NUMBER(18,6)", "oms.execution", "Fill price"],
    ["notional", "NUMBER(18,4)", "Derived", "fill_qty * fill_px"],
    ["commission", "NUMBER(18,4)", "oms.execution", "Fee"],
    ["venue", "STRING", "oms.execution", "NYSE/NASDAQ/ARCA/BATS/DARK"],
    ["fill_ts", "TIMESTAMP_NTZ", "oms.execution", "Exchange fill time"],
]
add_rows(ws3, cols)
style_header(ws3, 4)
for col, w in zip("ABCD", [18, 18, 22, 28]):
    ws3.column_dimensions[col].width = w

out = "/home/workdir/artifacts/aether-trade/docs/Aether_Trade_Data_Dictionary.xlsx"
wb.save(out)
print("saved", out)
